from __future__ import annotations

import argparse
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List
from xml.sax.saxutils import escape

from openpyxl import load_workbook


REPORT_HEADERS_CA = [
    "Ingredient Name",
    "CAS No.",
    "Concentration (%)",
    "INCI Name",
    "Conclusion",
]

REPORT_HEADERS_US = [
    "Ingredient Name",
    "CAS No.",
    "Concentration (%)",
    "INCI/Color additive Name",
    "Conclusion",
]


CIR_FINDING_TEXT = {
    "S": "Safe in the present practices of use and concentration.",
    "SQ": "safe for use in cosmetics, with qualifications;",
    "I": "the available data are insufficient to support safety;",
    "Z": "the available data are insufficient to support safety, but the ingredient is not used;",
    "U": "the ingredient is unsafe for use in cosmetics;",
    "UNS": "ingredients for which the data are insufficient and their use in cosmetics is not supported",
}

POSITIVE_CIR_FINDINGS = {"S", "SQ"}
NEGATIVE_CIR_FINDINGS = {"I", "Z", "U", "UNS"}

FRAGRANCE_KEYWORDS = ("FRAGRANCE", "PARFUM", "AROMA")
COLOR_ADDITIVE_KEYWORDS = (
    "FD&C",
    "FD & C",
    "D&C",
    "D & C",
    "EXT. D&C",
    "EXT D&C",
    " LAKE",
    "CI ",
)
COLOR_ADDITIVE_COLOR_WORDS = ("YELLOW", "RED", "BLUE", "GREEN", "VIOLET", "ORANGE", "BLACK", "BROWN")

FRAGRANCE_PREPARED_REMARK = (
    'Fragrance is a group name. The concentration of the fragrance in the formulation complies with the limit '
    'which indicated in the submitted 48h IFRA certificate of PL-247(G) Bee fragrance '
    '(Class 9.A maximum dosage 41.98%) submitted by the applicant. '
    'There is a typo in the submitted formulation. The correct name should be "fragrance".'
)

COLOR_ADDITIVE_PREPARED_REMARK = (
    "The color additive is listed in the 21 CFR 74 LISTING OF COLOR ADDITIVES SUBJECT TO CERTIFICATION. "
    "All batches of color additives listed in 21 CFR 74 shall be certified in accordance with regulations in part 80 of this chapter."
)
REPORT_DISCLAIMER = (
    "TÜV SÜD provides its services in a knowledge capacity only and offers no legal opinion(s) herein. "
    "The service is according to company’s know-how and on publicly available sources at the time the services were supplied."
)

PART_73_FULL_NAME = "21 CFR Part 73-Listing of Color Additives Exempt From Certification"
PART_74_FULL_NAME = "21 CFR Part 74-Listing of Color Additives Subject to Certification"
CA_PROHIBITED_FULL_NAME = "Health Canada Cosmetic Ingredient Hotlist-Prohibited"
CA_RESTRICTED_FULL_NAME = "Health Canada Cosmetic Ingredient Hotlist-Restricted"
MANUAL_REVIEW_RED = "C00000"


DOCX_CONTENT_TYPES = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
  <Override PartName="/word/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>
  <Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
  <Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>
</Types>
"""


DOCX_ROOT_RELS = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>
</Relationships>
"""


DOCX_DOCUMENT_RELS = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>
"""


DOCX_STYLES = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:docDefaults>
    <w:rPrDefault>
      <w:rPr>
        <w:rFonts w:ascii="Calibri" w:hAnsi="Calibri" w:eastAsia="Calibri" w:cs="Calibri"/>
        <w:sz w:val="20"/>
        <w:szCs w:val="20"/>
      </w:rPr>
    </w:rPrDefault>
    <w:pPrDefault>
      <w:pPr>
        <w:spacing w:after="120"/>
      </w:pPr>
    </w:pPrDefault>
  </w:docDefaults>
</w:styles>
"""


DOCX_APP_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties"
            xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
  <Application>Microsoft Office Word</Application>
  <DocSecurity>0</DocSecurity>
  <ScaleCrop>false</ScaleCrop>
  <Company></Company>
  <LinksUpToDate>false</LinksUpToDate>
  <SharedDoc>false</SharedDoc>
  <HyperlinksChanged>false</HyperlinksChanged>
  <AppVersion>16.0000</AppVersion>
</Properties>
"""


@dataclass
class FormulaReportData:
    formula_code: str
    formula_name: str
    product_family: str
    sheet_title: str
    rows: List[Dict[str, Any]]


def project_root() -> Path:
    return Path(__file__).resolve().parent


def output_dir(root: Path) -> Path:
    out = root / "output"
    out.mkdir(parents=True, exist_ok=True)
    return out


def find_latest_output_workbook(root: Path) -> Path:
    candidates = sorted(
        output_dir(root).glob("Cosmetic_Compliance_Output_*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError("No output workbook found under ./output.")
    return candidates[0]


def safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_text(value: Any) -> str:
    return " ".join(safe_str(value).lower().split())


def normalized_upper_text(*values: Any) -> str:
    return " ".join(safe_str(v).upper() for v in values if safe_str(v))


def format_concentration(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        text = f"{value:.6f}".rstrip("0").rstrip(".")
        return text or "0"
    return safe_str(value)


def get_inci_or_color_name(rec: Dict[str, Any]) -> str:
    if safe_str(rec.get("US Color Additive Trigger")).upper() == "Y":
        matched = safe_str(rec.get("US Color Matched Name"))
        if matched:
            return matched
    return safe_str(rec.get("Normalized / Screening Key Name")) or safe_str(rec.get("Raw INCI Name"))


def get_ca_inci_name(rec: Dict[str, Any]) -> str:
    return safe_str(rec.get("Normalized / Screening Key Name")) or safe_str(rec.get("Raw INCI Name"))


def is_fragrance_record(rec: Dict[str, Any]) -> bool:
    text = normalized_upper_text(rec.get("Raw INCI Name"), rec.get("Normalized / Screening Key Name"))
    return any(keyword in text for keyword in FRAGRANCE_KEYWORDS)


def is_color_additive_record(rec: Dict[str, Any]) -> bool:
    if safe_str(rec.get("US Color Additive Trigger")).upper() == "Y":
        return True
    text = normalized_upper_text(rec.get("Raw INCI Name"), rec.get("Normalized / Screening Key Name"))
    if any(keyword in text for keyword in COLOR_ADDITIVE_KEYWORDS):
        return True
    return any(f"{word} " in text for word in COLOR_ADDITIVE_COLOR_WORDS)


def is_color_additive_manual_review(rec: Dict[str, Any]) -> bool:
    return safe_str(rec.get("US Color Additive Trigger")).upper() == "Y"


def sanitize_filename(text: str) -> str:
    invalid = '<>:"/\\|?*'
    cleaned = "".join("_" if ch in invalid else ch for ch in safe_str(text))
    return cleaned.strip().strip(".") or "report"


def normalize_sheet_title(title: str) -> str:
    return title.strip()


def is_formula_sheet(ws) -> bool:
    title = normalize_sheet_title(ws.title)
    return title.startswith("03_") and title != "03_筛查输出模板" and ws.sheet_state == "visible"


def load_formula_reports(workbook_path: Path) -> List[FormulaReportData]:
    wb = load_workbook(workbook_path, data_only=True)
    reports: List[FormulaReportData] = []
    for ws in wb.worksheets:
        if not is_formula_sheet(ws):
            continue
        headers = [safe_str(ws.cell(2, c).value) for c in range(1, ws.max_column + 1)]
        rows: List[Dict[str, Any]] = []
        for r in range(3, ws.max_row + 1):
            values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if not any(v not in (None, "") for v in values):
                continue
            record = {headers[idx]: values[idx] for idx in range(len(headers)) if headers[idx]}
            if not safe_str(record.get("Raw INCI Name")):
                continue
            rows.append(record)
        if not rows:
            continue
        first = rows[0]
        reports.append(
            FormulaReportData(
                formula_code=safe_str(first.get("Formula Code")) or ws.title.removeprefix("03_"),
                formula_name=safe_str(first.get("Formula Name")) or safe_str(first.get("Formula Code")) or ws.title.removeprefix("03_"),
                product_family=safe_str(first.get("Product Family")),
                sheet_title=ws.title,
                rows=rows,
            )
        )
    return reports


def select_formula_reports(reports: List[FormulaReportData], formula_selector: str | None, include_all: bool) -> List[FormulaReportData]:
    if include_all:
        return reports
    if not formula_selector:
        raise ValueError("Please provide --formula for a sample report, or use --all.")
    selector = formula_selector.strip().casefold()
    selected = [
        report
        for report in reports
        if report.formula_code.casefold() == selector
        or report.formula_name.casefold() == selector
        or report.sheet_title.casefold() == f"03_{selector}"
    ]
    if not selected:
        raise ValueError(f"Formula '{formula_selector}' was not found in the workbook.")
    return selected


def join_parts(parts: Iterable[str]) -> str:
    seen: set[str] = set()
    cleaned: List[str] = []
    for part in parts:
        text = safe_str(part)
        if not text:
            continue
        if text in seen:
            continue
        seen.add(text)
        cleaned.append(text)
    return "\n".join(cleaned)


def build_ca_entry(rec: Dict[str, Any]) -> tuple[str, str, bool, str]:
    prohibited = safe_str(rec.get("CA Prohibited Hit")).upper()
    restricted = safe_str(rec.get("CA Restricted Hit")).upper()
    manual = safe_str(rec.get("Manual Review?")).upper() in {"Y", "MAYBE"}

    if is_fragrance_record(rec):
        remark = (
            "The substance is fragrance, which is a group name rather than a single defined ingredient. "
            "Manual review is required against the submitted IFRA documentation and label disclosure requirements."
        )
        return "Manual Review", remark, True, "CA_FRAGRANCE_MANUAL_REVIEW"

    if prohibited == "Y":
        evidence = safe_str(rec.get("CA Prohibited Evidence"))
        remark = f"The substance was found in {CA_PROHIBITED_FULL_NAME}. "
        if evidence:
            remark += f"Evidence: {evidence}. "
        remark += "This substance is prohibited for cosmetic use in Canada."
        if manual:
            remark += " Manual review required."
        return "Not Compliance", remark, manual, f"CA_PROHIBITED|{normalize_text(evidence)}"

    if restricted == "Y":
        matched_entry = safe_str(rec.get("CA Restricted Matched Ingredient"))
        primary_entry = matched_entry.split("|")[0].strip() if matched_entry else ""
        conditions = safe_str(rec.get("Conditions of Use / Allowed Body Site"))
        max_conc = safe_str(rec.get("Max Conc. Permitted"))
        warning = safe_str(rec.get("Warning Statement"))
        parts: List[str] = [f"The substance was found in {CA_RESTRICTED_FULL_NAME}."]
        if primary_entry:
            parts.append(f"Relevant Hotlist entry: {primary_entry}.")
        if conditions:
            parts.append(f"Conditions of use / allowed body site: {conditions}.")
        if max_conc:
            parts.append(f"Maximum permitted concentration: {max_conc}.")
        if warning:
            parts.append(f"Warning statement: {warning}.")
        parts.append("Manual review is required to confirm product type, body site, concentration, and labeling.")
        remark = " ".join(parts)
        remark_key = "|".join(
            [
                "CA_RESTRICTED",
                normalize_text(primary_entry),
                normalize_text(max_conc),
                normalize_text(warning),
            ]
        )
        return "Manual Review", remark, True, remark_key

    return (
        "Compliance",
        "The substance was not found in the Health Canada Cosmetic Ingredient Hotlist prohibited or restricted lists.",
        False,
        "CA_NO_HIT",
    )


def build_ca_prepared_remarks(report: FormulaReportData) -> List[Dict[str, Any]]:
    prepared: List[Dict[str, Any]] = []
    if any(safe_str(rec.get("CA Restricted Hit")).upper() == "Y" for rec in report.rows):
        prepared.append(
            {
                "text": "Restricted entries in the Health Canada Cosmetic Ingredient Hotlist should be reviewed against product type, body site, concentration, and warning statement requirements before release.",
                "color": MANUAL_REVIEW_RED,
            }
        )

    fragrance_rows = [rec for rec in report.rows if is_fragrance_record(rec)]
    if fragrance_rows:
        fragrance_names = []
        for rec in fragrance_rows:
            raw_name = safe_str(rec.get("Raw INCI Name"))
            if raw_name and raw_name.upper() not in {"FRAGRANCE", "PARFUM", "AROMA"} and raw_name not in fragrance_names:
                fragrance_names.append(raw_name)
        fragrance_label = fragrance_names[0] if fragrance_names else "[fragrance name from the submitted IFRA certificate]"
        prepared.append(
            {
                "text": (
                    f'Fragrance is a group name. The concentration of the fragrance in the formulation complies with the limit indicated in the submitted 51th IFRA certificate of {fragrance_label} submitted by the applicant.'
                ),
                "color": MANUAL_REVIEW_RED,
            }
        )
        prepared.append(
            {
                "text": "If allergenic fragrance ingredients exceed 0.01% in rinse-off products or 0.001% in leave-on products, the specific allergenic fragrance ingredients should be declared on the label.",
                "color": MANUAL_REVIEW_RED,
            }
        )
    return prepared


def build_ca_report_rows(report: FormulaReportData) -> tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    rows: List[Dict[str, Any]] = []
    remark_numbers: Dict[str, int] = {}
    ordered_remarks: List[Dict[str, Any]] = []
    remark_color_flags: Dict[str, bool] = {}

    for rec in report.rows:
        status, remark_text, needs_red, remark_key = build_ca_entry(rec)
        key = remark_key or remark_text
        if key not in remark_numbers:
            remark_numbers[key] = len(ordered_remarks) + 1
            ordered_remarks.append({"text": remark_text, "color": MANUAL_REVIEW_RED if needs_red else None})
            remark_color_flags[key] = needs_red
        elif needs_red and not remark_color_flags.get(key):
            remark_color_flags[key] = True
            ordered_remarks[remark_numbers[key] - 1]["color"] = MANUAL_REVIEW_RED
        remark_number = remark_numbers[key]
        conclusion = f"{status} {remark_number})"
        if needs_red and status != "Manual Review":
            conclusion += " / Manual Review"
        rows.append(
            {
                "cells": [
                    safe_str(rec.get("Raw INCI Name")),
                    safe_str(rec.get("Raw CAS No")),
                    format_concentration(rec.get("Input W/W %")),
                    get_ca_inci_name(rec),
                    conclusion,
                ],
                "color": MANUAL_REVIEW_RED if needs_red else None,
            }
        )

    remarks = [
        {"text": f"{idx}) {item['text']}", "color": item.get("color")}
        for idx, item in enumerate(ordered_remarks, start=1)
    ]
    remarks.append({"text": REPORT_DISCLAIMER, "color": None})
    return rows, remarks


def extract_cir_finding_codes(raw_finding: Any) -> List[str]:
    text = safe_str(raw_finding).upper().replace("†", "")
    if not text:
        return []
    for sep in ["|", ";", "/", ","]:
        text = text.replace(sep, " ")
    seen: set[str] = set()
    codes: List[str] = []
    for token in text.split():
        if token in CIR_FINDING_TEXT and token not in seen:
            seen.add(token)
            codes.append(token)
    return codes


def classify_cir_status(codes: List[str]) -> str:
    if not codes:
        return "Not Compliance"
    has_positive = any(code in POSITIVE_CIR_FINDINGS for code in codes)
    has_negative = any(code in NEGATIVE_CIR_FINDINGS for code in codes)
    if has_positive and has_negative:
        return "Manual Review"
    if has_positive:
        return "Compliance"
    if has_negative:
        return "Not Compliance"
    return "Not Compliance"


def build_us_color_entry(rec: Dict[str, Any]) -> tuple[str, str, bool]:
    category = safe_str(rec.get("US Color Category"))
    match_status = safe_str(rec.get("US Color Match Status")).upper()
    matched_name = safe_str(rec.get("US Color Matched Name")) or get_inci_or_color_name(rec)
    uses = safe_str(rec.get("US Color Uses and Restrictions"))

    if "73 Subpart C" in category:
        restriction = (uses or "[blank in source]").rstrip().rstrip(".")
        remark = (
            f"The substance was found as {matched_name} in {PART_73_FULL_NAME}. "
            f"It does not require certification. Use restriction is {restriction}. Manual review required."
        )
        return "Compliance", remark, True

    if "74 Subpart C" in category:
        restriction = (uses or "[blank in source]").rstrip().rstrip(".")
        remark = (
            f"The substance was found as {matched_name} in {PART_74_FULL_NAME}. "
            f"It requires certification. Use restriction is {restriction}. Manual review required."
        )
        return "Compliance", remark, True

    if match_status == "MATCHED":
        detail = f" Current database classification: {category}." if category else ""
        if uses:
            detail += f" Recorded use restriction: {uses}."
        remark = (
            "The substance was not found in the CFR color additive lists permitted for cosmetic use."
            f"{detail} Manual review required."
        )
        return "Not Compliance", remark, True

    return (
        "Not Compliance",
        "The substance was not found in the CFR color additive lists permitted for cosmetic use. Manual review required.",
        True,
    )


def build_us_entry(rec: Dict[str, Any]) -> tuple[str, str, bool]:
    if is_fragrance_record(rec):
        return "Manual Review", "The substance is fragrance. Fragrance review workflow is not implemented in this version and requires manual review.", False

    if is_color_additive_record(rec):
        return build_us_color_entry(rec)

    cfr = safe_str(rec.get("CFR Hit")).upper()
    cir = safe_str(rec.get("CIR Hit")).upper()

    if cfr == "Y":
        return "Not Compliance", "The substance was found in the CFR prohibited substances list.", False

    if cir != "Y":
        return (
            "Compliance",
            "The substance was not found in the CFR prohibited and restricted substances list or the CIR Quick Reference Table.",
            False,
        )

    raw_finding = safe_str(rec.get("Finding"))
    source_conclusion = safe_str(rec.get("Conclusion"))
    codes = extract_cir_finding_codes(raw_finding)

    status = classify_cir_status(codes)

    if codes and any(code in POSITIVE_CIR_FINDINGS for code in codes) and any(code in NEGATIVE_CIR_FINDINGS for code in codes):
        remark = (
            "The substance was found in the CIR Quick Reference Table with mixed finding codes "
            f"'{raw_finding}', including both supportive and insufficient/negative safety signals. Manual review is required."
        )
        if source_conclusion:
            remark += f" Conclusion in the Quick Reference Table: {source_conclusion}"
        return status, remark, False

    if codes == ["S"]:
        return status, "The substance was found in the CIR Quick Reference Table, and is Safe in the present practices of use and concentration.", False

    if codes == ["SQ"]:
        remark = (
            "The substance was found in the CIR Quick Reference Table, "
            "and is safe for use in cosmetics, with qualifications."
        )
        if source_conclusion:
            remark += f" Conclusion in the Quick Reference Table: {source_conclusion}"
        return status, remark, False

    if codes == ["I"]:
        return status, "The substance was found in the CIR Quick Reference Table, and the available data are insufficient to support safety.", False

    if codes == ["Z"]:
        return status, "The substance was found in the CIR Quick Reference Table, and the available data are insufficient to support safety, but the ingredient is not used.", False

    if codes == ["U"]:
        return status, "The substance was found in the CIR Quick Reference Table, and the ingredient is unsafe for use in cosmetics.", False

    if codes == ["UNS"]:
        return status, "The substance was found in the CIR Quick Reference Table, and ingredients for which the data are insufficient and their use in cosmetics is not supported.", False

    if raw_finding:
        remark = (
            "The substance was found in the CIR Quick Reference Table with finding "
            f"'{raw_finding}'."
        )
        if source_conclusion:
            remark += f" Conclusion in the Quick Reference Table: {source_conclusion}"
        return status, remark, False

    return "Not Compliance", (
        "The substance was found in the CIR Quick Reference Table, "
        "but the finding code is blank and needs manual confirmation."
    ), False


def build_us_prepared_remarks(report: FormulaReportData) -> List[str]:
    prepared: List[str] = []
    if any(is_fragrance_record(rec) for rec in report.rows):
        prepared.append(f"Fragrance: {FRAGRANCE_PREPARED_REMARK}")
    return prepared


def build_us_report_rows(report: FormulaReportData) -> tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    rows: List[Dict[str, Any]] = []
    remark_numbers: Dict[str, int] = {}
    ordered_remarks: List[Dict[str, Any]] = []
    remark_color_flags: Dict[str, bool] = {}

    for rec in report.rows:
        status, remark_text, is_color_review = build_us_entry(rec)
        if remark_text not in remark_numbers:
            remark_numbers[remark_text] = len(ordered_remarks) + 1
            ordered_remarks.append({"text": remark_text, "color": MANUAL_REVIEW_RED if is_color_review else None})
            remark_color_flags[remark_text] = is_color_review
        elif is_color_review and not remark_color_flags.get(remark_text):
            remark_color_flags[remark_text] = True
            ordered_remarks[remark_numbers[remark_text] - 1]["color"] = MANUAL_REVIEW_RED
        remark_number = remark_numbers[remark_text]
        conclusion = f"{status} {remark_number})"
        if is_color_review:
            conclusion += " / Manual Review"
        rows.append(
            {
                "cells": [
                    safe_str(rec.get("Raw INCI Name")),
                    safe_str(rec.get("Raw CAS No")),
                    format_concentration(rec.get("Input W/W %")),
                    get_inci_or_color_name(rec),
                    conclusion,
                ],
                "color": MANUAL_REVIEW_RED if is_color_review else None,
            }
        )

    remarks = [
        {"text": f"{idx}) {item['text']}", "color": item.get("color")}
        for idx, item in enumerate(ordered_remarks, start=1)
    ]
    remarks.append({"text": REPORT_DISCLAIMER, "color": None})
    return rows, remarks


def build_report_rows(report: FormulaReportData, region: str) -> List[List[str]]:
    if region == "CA":
        raise ValueError("CA report rows should be built via build_ca_report_rows().")
    if region == "US":
        raise ValueError("US report rows should be built via build_us_report_rows().")
    raise ValueError(f"Unsupported region: {region}")


def build_report_notes(report: FormulaReportData, region: str, workbook_name: str) -> List[str]:
    if region == "CA":
        return []

    return []


def xml_text(text: str) -> str:
    return f'<w:t xml:space="preserve">{escape(text)}</w:t>'


def build_run_xml(text: str, *, bold: bool = False, size: int = 20, color: str | None = None) -> str:
    run_props: List[str] = []
    if bold:
        run_props.extend(["<w:b/>", "<w:bCs/>"])
    if size:
        run_props.append(f'<w:sz w:val="{size}"/>')
        run_props.append(f'<w:szCs w:val="{size}"/>')
    if color:
        run_props.append(f'<w:color w:val="{color}"/>')
    props_xml = f"<w:rPr>{''.join(run_props)}</w:rPr>" if run_props else ""

    lines = text.split("\n") if text else [""]
    chunks: List[str] = []
    for idx, line in enumerate(lines):
        if idx > 0:
            chunks.append("<w:br/>")
        chunks.append(xml_text(line))
    return f"<w:r>{props_xml}{''.join(chunks)}</w:r>"


def build_paragraph_xml(
    text: str,
    *,
    bold: bool = False,
    size: int = 20,
    color: str | None = None,
    align: str | None = None,
    space_after: int = 120,
    space_before: int = 0,
) -> str:
    para_props: List[str] = []
    if align:
        para_props.append(f'<w:jc w:val="{align}"/>')
    para_props.append(f'<w:spacing w:before="{space_before}" w:after="{space_after}"/>')
    props_xml = f"<w:pPr>{''.join(para_props)}</w:pPr>"
    return f"<w:p>{props_xml}{build_run_xml(text, bold=bold, size=size, color=color)}</w:p>"


def build_cell_xml(text: str, *, width: int, header: bool = False, color: str | None = None) -> str:
    tc_props = [f'<w:tcW w:w="{width}" w:type="dxa"/>']
    if header:
        tc_props.append('<w:shd w:val="clear" w:color="auto" w:fill="D9E2F3"/>')
    props_xml = f"<w:tcPr>{''.join(tc_props)}</w:tcPr>"
    para_xml = build_paragraph_xml(text, bold=header, size=20, color=color, space_after=0)
    return f"<w:tc>{props_xml}{para_xml}</w:tc>"


def build_table_widths(headers: List[str]) -> List[int]:
    if headers == REPORT_HEADERS_US:
        return [2400, 2000, 1400, 2800, 2200]
    if headers == REPORT_HEADERS_CA:
        return [2200, 1900, 1300, 2600, 2200]
    return [2200, 1900, 1300, 2200, 5200]


def build_table_xml(headers: List[str], rows: List[Any]) -> str:
    widths = build_table_widths(headers)
    grid_xml = "".join(f'<w:gridCol w:w="{width}"/>' for width in widths)
    borders_xml = """
<w:tblBorders>
  <w:top w:val="single" w:sz="8" w:space="0" w:color="808080"/>
  <w:left w:val="single" w:sz="8" w:space="0" w:color="808080"/>
  <w:bottom w:val="single" w:sz="8" w:space="0" w:color="808080"/>
  <w:right w:val="single" w:sz="8" w:space="0" w:color="808080"/>
  <w:insideH w:val="single" w:sz="4" w:space="0" w:color="BFBFBF"/>
  <w:insideV w:val="single" w:sz="4" w:space="0" w:color="BFBFBF"/>
</w:tblBorders>
""".strip()
    header_row = "<w:tr>" + "".join(
        build_cell_xml(text, width=widths[idx], header=True) for idx, text in enumerate(headers)
    ) + "</w:tr>"

    body_rows: List[str] = []
    for row in rows:
        row_cells = row["cells"] if isinstance(row, dict) else row
        row_color = row.get("color") if isinstance(row, dict) else None
        cells = "".join(
            build_cell_xml(safe_str(row_cells[idx]), width=widths[idx], header=False, color=row_color)
            for idx in range(len(headers))
        )
        body_rows.append(f"<w:tr>{cells}</w:tr>")

    return (
        "<w:tbl>"
        "<w:tblPr>"
        '<w:tblW w:w="0" w:type="auto"/>'
        f"{borders_xml}"
        "</w:tblPr>"
        f"<w:tblGrid>{grid_xml}</w:tblGrid>"
        f"{header_row}"
        f"{''.join(body_rows)}"
        "</w:tbl>"
    )


def build_document_xml(
    title: str,
    subtitle_lines: List[str],
    headers: List[str],
    rows: List[Any],
    notes: List[Any],
    notes_heading: str | None,
    extra_sections: List[tuple[str, List[str]]] | None = None,
) -> str:
    body_parts: List[str] = [
        build_paragraph_xml(title, bold=True, size=32, align="center", space_after=240),
    ]
    for line in subtitle_lines:
        body_parts.append(build_paragraph_xml(line, size=20, space_after=40))
    body_parts.append(build_paragraph_xml("", size=20, space_after=80))
    body_parts.append(build_table_xml(headers, rows))
    if notes and notes_heading:
        body_parts.append(build_paragraph_xml("", size=20, space_after=80))
        body_parts.append(build_paragraph_xml(notes_heading, bold=True, size=24, space_after=120))
        for note in notes:
            if isinstance(note, dict):
                body_parts.append(build_paragraph_xml(note["text"], size=20, color=note.get("color"), space_after=40))
            else:
                body_parts.append(build_paragraph_xml(note, size=20, space_after=40))
    for heading, lines in extra_sections or []:
        if not lines:
            continue
        body_parts.append(build_paragraph_xml("", size=20, space_after=80))
        body_parts.append(build_paragraph_xml(heading, bold=True, size=24, space_after=120))
        for line in lines:
            if isinstance(line, dict):
                body_parts.append(build_paragraph_xml(line["text"], size=20, color=line.get("color"), space_after=40))
            else:
                body_parts.append(build_paragraph_xml(line, size=20, space_after=40))
    body_parts.append(
        '<w:sectPr>'
        '<w:pgSz w:w="12240" w:h="15840"/>'
        '<w:pgMar w:top="1440" w:right="720" w:bottom="1440" w:left="720" w:header="708" w:footer="708" w:gutter="0"/>'
        "</w:sectPr>"
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:wpc="http://schemas.microsoft.com/office/word/2010/wordprocessingCanvas" '
        'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
        'xmlns:o="urn:schemas-microsoft-com:office:office" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        'xmlns:m="http://schemas.openxmlformats.org/officeDocument/2006/math" '
        'xmlns:v="urn:schemas-microsoft-com:vml" '
        'xmlns:wp14="http://schemas.microsoft.com/office/word/2010/wordprocessingDrawing" '
        'xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing" '
        'xmlns:w10="urn:schemas-microsoft-com:office:word" '
        'xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main" '
        'xmlns:w14="http://schemas.microsoft.com/office/word/2010/wordml" '
        'xmlns:wpg="http://schemas.microsoft.com/office/word/2010/wordprocessingGroup" '
        'xmlns:wpi="http://schemas.microsoft.com/office/word/2010/wordprocessingInk" '
        'xmlns:wne="http://schemas.microsoft.com/office/word/2006/wordml" '
        'xmlns:wps="http://schemas.microsoft.com/office/word/2010/wordprocessingShape" '
        'mc:Ignorable="w14 wp14">'
        f"<w:body>{''.join(body_parts)}</w:body>"
        "</w:document>"
    )


def build_core_xml(title: str) -> str:
    created = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties"
                   xmlns:dc="http://purl.org/dc/elements/1.1/"
                   xmlns:dcterms="http://purl.org/dc/terms/"
                   xmlns:dcmitype="http://purl.org/dc/dcmitype/"
                   xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <dc:title>{escape(title)}</dc:title>
  <dc:creator>Codex</dc:creator>
  <cp:lastModifiedBy>Codex</cp:lastModifiedBy>
  <dcterms:created xsi:type="dcterms:W3CDTF">{created}</dcterms:created>
  <dcterms:modified xsi:type="dcterms:W3CDTF">{created}</dcterms:modified>
</cp:coreProperties>
"""


def write_docx(
    output_path: Path,
    title: str,
    subtitle_lines: List[str],
    headers: List[str],
    rows: List[Any],
    notes: List[Any],
    notes_heading: str | None,
    extra_sections: List[tuple[str, List[str]]] | None = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    document_xml = build_document_xml(title, subtitle_lines, headers, rows, notes, notes_heading, extra_sections)
    core_xml = build_core_xml(title)
    with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", DOCX_CONTENT_TYPES)
        zf.writestr("_rels/.rels", DOCX_ROOT_RELS)
        zf.writestr("docProps/app.xml", DOCX_APP_XML)
        zf.writestr("docProps/core.xml", core_xml)
        zf.writestr("word/document.xml", document_xml)
        zf.writestr("word/styles.xml", DOCX_STYLES)
        zf.writestr("word/_rels/document.xml.rels", DOCX_DOCUMENT_RELS)


def build_subtitle_lines(report: FormulaReportData, workbook_path: Path, region: str) -> List[str]:
    generated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return [
        f"Region: {region}",
        f"Formula Code: {report.formula_code}",
        f"Formula Name: {report.formula_name}",
        f"Product Family: {report.product_family or '[blank]'}",
        f"Source Workbook: {workbook_path.name}",
        f"Source Sheet: {report.sheet_title}",
        f"Generated At: {generated}",
    ]


def generate_formula_reports(workbook_path: Path, report: FormulaReportData, base_output_dir: Path) -> List[Path]:
    formula_dir = base_output_dir / sanitize_filename(report.formula_code)
    ca_title = f"CA Compliance Screening Report - {report.formula_code}"
    us_title = f"US Compliance Screening Report - {report.formula_code}"

    ca_path = formula_dir / f"{sanitize_filename(report.formula_code)}_CA_Report_v2.docx"
    us_path = formula_dir / f"{sanitize_filename(report.formula_code)}_US_Report_v2.docx"

    ca_rows, ca_remarks = build_ca_report_rows(report)
    ca_prepared_remarks = build_ca_prepared_remarks(report)
    us_rows, us_remarks = build_us_report_rows(report)
    us_prepared_remarks = build_us_prepared_remarks(report)

    write_docx(
        ca_path,
        ca_title,
        build_subtitle_lines(report, workbook_path, "CA"),
        REPORT_HEADERS_CA,
        ca_rows,
        ca_remarks,
        "Remarks",
        [("Prepared Remarks", ca_prepared_remarks)],
    )
    write_docx(
        us_path,
        us_title,
        build_subtitle_lines(report, workbook_path, "US"),
        REPORT_HEADERS_US,
        us_rows,
        us_remarks,
        "Remarks",
        [("Prepared Remarks", us_prepared_remarks)],
    )
    return [ca_path, us_path]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate CA and US Word reports from the screening output workbook.")
    parser.add_argument("--workbook", type=Path, help="Path to a Cosmetic_Compliance_Output workbook.")
    parser.add_argument("--formula", help="Formula code or formula name for a sample report.")
    parser.add_argument("--all", action="store_true", help="Generate reports for all formula sheets in the workbook.")
    parser.add_argument("--output-dir", type=Path, help="Directory for generated Word reports.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    root = project_root()
    workbook_path = args.workbook.resolve() if args.workbook else find_latest_output_workbook(root)
    reports = load_formula_reports(workbook_path)
    selected = select_formula_reports(reports, args.formula, args.all)

    default_output_dir = output_dir(root) / f"word_reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    report_output_dir = args.output_dir.resolve() if args.output_dir else default_output_dir
    generated_paths: List[Path] = []
    for report in selected:
        generated_paths.extend(generate_formula_reports(workbook_path, report, report_output_dir))

    print(f"Workbook: {workbook_path}")
    print(f"Generated formula reports: {len(selected)}")
    print(f"Output directory: {report_output_dir}")
    for path in generated_paths:
        print(path)


if __name__ == "__main__":
    main()
