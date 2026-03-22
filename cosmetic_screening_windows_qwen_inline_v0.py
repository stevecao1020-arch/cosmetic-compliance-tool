from __future__ import annotations

import json
import os
import re
import shutil
import sys
import time
import traceback
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from urllib import request, error

try:
    import pandas as pd
except Exception:
    print("[ERROR] pandas 未安装。请先执行: pip install pandas openpyxl rapidfuzz")
    raise

try:
    from openpyxl import load_workbook
except Exception:
    print("[ERROR] openpyxl 未安装。请先执行: pip install openpyxl")
    raise

try:
    from rapidfuzz import fuzz
except Exception:
    from difflib import SequenceMatcher

    class _CompatFuzz:
        @staticmethod
        def ratio(a: str, b: str) -> int:
            return int(round(SequenceMatcher(None, a, b).ratio() * 100))

    fuzz = _CompatFuzz()

# ============================================================
# Windows 项目结构
# 桌面\CA&US成分审核
# ├─ input
# ├─ database
# ├─ template
# ├─ output
# └─ logs
# ============================================================
APP_ROOT_NAME = "CA&US成分审核"
INPUT_DIR_NAME = "input"
DATABASE_DIR_NAME = "database"
TEMPLATE_DIR_NAME = "template"
OUTPUT_DIR_NAME = "output"
LOG_DIR_NAME = "logs"

DB_FILE_HINTS = {
    "ca_prohibited": ["ca", "hotlist", "prohibited"],
    "ca_restricted": ["ca", "hotlist", "restricted"],
    "cir": ["cir", "quick", "reference"],
    "cfr_banned": ["cfr", "禁用"],
}
TEMPLATE_FILE_HINTS = ["output", "template"]
CLIENT_FILE_SUFFIXES = {".xlsx", ".xlsm", ".xls"}

AI_MODE = True
AI_FUZZY_THRESHOLD = 82
MAX_FUZZY_CANDIDATES = 5
AI_ENABLED_ONLY_FOR_LOW_CONFIDENCE = True
API_TIMEOUT_SECONDS = 90

# ============================================================
# 千问 API 直接写在这里（双击 .py 即可运行）
# 你只需要填写 DASHSCOPE_API_KEY
# 常用可选项：
# - 国内站: https://dashscope.aliyuncs.com/compatible-mode/v1
# - 国际站: https://dashscope-intl.aliyuncs.com/compatible-mode/v1
# ============================================================
DASHSCOPE_API_KEY = ""   # <<< 在这里填写你的千问 API Key
QWEN_BASE_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1"
QWEN_MODEL = "qwen-plus"

# ============================================================
# 基础工具
# ============================================================
def normalize_text(text: object) -> str:
    if text is None:
        return ""
    s = str(text).strip().lower()
    s = s.replace("（", "(").replace("）", ")")
    s = s.replace("–", "-").replace("—", "-")
    s = s.replace("／", "/")
    s = s.replace("\u00a0", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def normalize_name_key(text: object) -> str:
    s = normalize_text(text)
    s = s.replace("&", " and ")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def flatten_name_key(text: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize_text(text))


def name_key_variants(text: object) -> List[str]:
    raw = normalize_text(text)
    variants: List[str] = []
    def add(v: str):
        nk = normalize_name_key(v)
        if nk and nk not in variants:
            variants.append(nk)
    add(raw)
    # remove parenthetical abbreviations / notes
    no_paren = re.sub(r"\([^)]*\)", " ", raw)
    add(no_paren)
    # split only on explicit alias separators; keep commas inside chemical names
    for seg in re.split(r"[;/\n]+", no_paren):
        add(seg)
    return variants


def exact_name_keys(text: object) -> List[str]:
    raw = normalize_text(text)
    variants: List[str] = []

    def add(v: str):
        nk = normalize_name_key(v)
        if nk and nk not in variants:
            variants.append(nk)

    add(raw)
    add(re.sub(r"\([^)]*\)", " ", raw))
    return variants


def cas_variants(text: object) -> List[str]:
    vals = []
    for p in split_multi_cas_value(text):
        nc = normalize_cas(p)
        if nc and nc not in vals:
            vals.append(nc)
    return vals


def split_multi_text_value(text: object) -> List[str]:
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return []
    s = str(text)
    s = s.replace("\u00a0", " ")
    parts = re.split(r"[;/\n]+", s)
    out = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        out.append(p)
    return out


def split_multi_cas_value(text: object) -> List[str]:
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return []
    s = str(text)
    s = s.replace("\u00a0", " ")
    parts = re.split(r"[;,/\n]+", s)
    out = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        out.append(p)
    return out


def split_multi_value(text: object) -> List[str]:
    return split_multi_text_value(text)


def normalize_cas(cas: object) -> str:
    s = normalize_text(cas)
    s = s.replace(" ", "")
    s = s.rstrip(".")
    return s


def minor_name_token_variant(a: str, b: str) -> bool:
    if not a or not b:
        return False
    if a == b:
        return True
    shorter, longer = (a, b) if len(a) <= len(b) else (b, a)
    if len(shorter) >= 6 and longer.startswith(shorter):
        suffix = longer[len(shorter):]
        if suffix in {"s", "es", "ed", "ing", "ion", "ions"}:
            return True
    if len(a) >= 6 and len(b) >= 6 and abs(len(a) - len(b)) <= 2:
        if fuzz.ratio(a, b) >= 92:
            return True
    return False


def low_confidence_name_reason(raw_name: str, candidate_name: str) -> str:
    raw_key = normalize_name_key(raw_name)
    candidate_key = normalize_name_key(candidate_name)
    if not raw_key or not candidate_key or raw_key == candidate_key:
        return ""
    if flatten_name_key(raw_key) == flatten_name_key(candidate_key):
        return "punctuation/spacing variation"

    raw_tokens = raw_key.split()
    candidate_tokens = candidate_key.split()
    if not raw_tokens or len(raw_tokens) != len(candidate_tokens):
        return ""
    if all(minor_name_token_variant(a, b) for a, b in zip(raw_tokens, candidate_tokens)):
        if any(a != b for a, b in zip(raw_tokens, candidate_tokens)):
            return "minor spelling/suffix variation"
    return ""


def row_match_evidence(row: dict, raw_name_keys: List[str], raw_cas_keys: List[str]) -> Optional[dict]:
    row_name_key = normalize_name_key(row.get("ingredient_name", ""))
    synonym_keys = row.get("synonym_list", []) or []
    row_cas_keys = row.get("cas_list", []) or []

    name_via_primary = bool(row_name_key and row_name_key in raw_name_keys)
    name_via_synonym = any(k in synonym_keys for k in raw_name_keys)
    name_match = name_via_primary or name_via_synonym
    cas_match = any(c in row_cas_keys for c in raw_cas_keys)
    row_has_cas = bool(row_cas_keys)
    raw_has_cas = bool(raw_cas_keys)

    if name_match and cas_match:
        category = "NAME_AND_CAS_MATCH"
    elif cas_match:
        category = "CAS_MATCH_ONLY"
    elif name_match and raw_has_cas and row_has_cas:
        category = "NAME_MATCH_CAS_MISMATCH"
    elif name_match:
        category = "NAME_MATCH_ONLY"
    else:
        return None

    return {
        "row": row,
        "category": category,
        "name_match": name_match,
        "name_via_primary": name_via_primary,
        "name_via_synonym": name_via_synonym,
        "cas_match": cas_match,
        "row_has_cas": row_has_cas,
        "source_db": row.get("db_source", ""),
    }


def best_row_match_evidence(rows: List[dict], raw_name_keys: List[str], raw_cas_keys: List[str]) -> Optional[dict]:
    category_rank = {
        "NAME_AND_CAS_MATCH": 0,
        "CAS_MATCH_ONLY": 1,
        "NAME_MATCH_CAS_MISMATCH": 2,
        "NAME_MATCH_ONLY": 3,
    }
    source_rank = {
        "CA_PROHIBITED": 0,
        "CA_RESTRICTED": 1,
        "CFR_BANNED": 2,
        "CIR": 3,
    }
    best = None
    best_rank = None
    for row in rows:
        evidence = row_match_evidence(row, raw_name_keys, raw_cas_keys)
        if not evidence:
            continue
        rank = (
            category_rank.get(evidence["category"], 99),
            source_rank.get(evidence["source_db"], 99),
            0 if evidence["name_via_primary"] else 1,
        )
        if best is None or rank < best_rank:
            best = evidence
            best_rank = rank
    return best


def desktop_dir() -> Path:
    return Path.home() / "Desktop"


def project_root() -> Path:
    return desktop_dir() / APP_ROOT_NAME


def ensure_project_structure(root: Path) -> Dict[str, Path]:
    paths = {
        "root": root,
        "input": root / INPUT_DIR_NAME,
        "database": root / DATABASE_DIR_NAME,
        "template": root / TEMPLATE_DIR_NAME,
        "output": root / OUTPUT_DIR_NAME,
        "logs": root / LOG_DIR_NAME,
    }
    for p in paths.values():
        p.mkdir(parents=True, exist_ok=True)
    return paths


def list_excel_files(folder: Path) -> List[Path]:
    if not folder.exists():
        return []
    return sorted([p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in CLIENT_FILE_SUFFIXES])


def find_file_by_keywords(folder: Path, keyword_groups: Dict[str, List[str]]) -> Dict[str, Optional[Path]]:
    files = list_excel_files(folder)
    result: Dict[str, Optional[Path]] = {k: None for k in keyword_groups}
    for key, keywords in keyword_groups.items():
        for f in files:
            name = normalize_text(f.name)
            if all(k in name for k in keywords):
                result[key] = f
                break
    return result


def find_template_file(folder: Path) -> Optional[Path]:
    files = list_excel_files(folder)
    for f in files:
        name = normalize_text(f.name)
        if all(k in name for k in TEMPLATE_FILE_HINTS):
            return f
    if len(files) == 1:
        return files[0]
    return None


def print_header(title: str) -> None:
    print("\n" + "=" * 72)
    print(title)
    print("=" * 72)


def safe_float(value: object) -> Optional[float]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        return float(value)
    except Exception:
        s = str(value).strip().replace("%", "")
        try:
            return float(s)
        except Exception:
            return None


def read_env_file(env_path: Path) -> None:
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


# ============================================================
# API / AI
# 支持 OpenAI 兼容接口：
# - Qwen / DashScope 兼容模式（优先）
# - OpenAI / 其他兼容网关
# 环境变量优先级：
# 1) DASHSCOPE_API_KEY
# 2) OPENAI_API_KEY
# ============================================================
@dataclass
class AIConfig:
    api_key: str
    base_url: str
    model: str
    enabled: bool


# 说明：保留环境变量兜底，但默认不依赖 .env

def load_ai_config(root: Path) -> AIConfig:
    read_env_file(root / ".env")

    # 优先级：代码顶部手填 > .env / 系统环境变量 > 默认值
    api_key = (
        (DASHSCOPE_API_KEY or "").strip()
        or os.getenv("DASHSCOPE_API_KEY", "").strip()
        or os.getenv("OPENAI_API_KEY", "").strip()
    )

    base_url = (
        (QWEN_BASE_URL or "").strip()
        or os.getenv("OPENAI_BASE_URL", "").strip()
        or os.getenv("DASHSCOPE_BASE_URL", "").strip()
        or "https://dashscope.aliyuncs.com/compatible-mode/v1"
    ).rstrip("/")

    model = (
        (QWEN_MODEL or "").strip()
        or os.getenv("OPENAI_MODEL", "").strip()
        or os.getenv("DASHSCOPE_MODEL", "").strip()
        or "qwen-plus"
    )

    enabled = bool(api_key and AI_MODE)
    return AIConfig(api_key=api_key, base_url=base_url, model=model, enabled=enabled)


def call_openai_compatible_json(ai: AIConfig, system_prompt: str, user_prompt: str, temperature: float = 0.1) -> dict:
    if not ai.enabled:
        raise RuntimeError("AI 未启用：请检查 DASHSCOPE_API_KEY / OPENAI_BASE_URL / OPENAI_MODEL")

    payload = {
        "model": ai.model,
        "temperature": temperature,
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
    }
    data = json.dumps(payload).encode("utf-8")
    req = request.Request(
        url=f"{ai.base_url}/chat/completions",
        data=data,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {ai.api_key}",
        },
        method="POST",
    )
    try:
        with request.urlopen(req, timeout=API_TIMEOUT_SECONDS) as resp:
            body = resp.read().decode("utf-8")
    except error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"API HTTP错误 {e.code}: {detail}")
    except Exception as e:
        raise RuntimeError(f"API 调用失败: {e}")

    obj = json.loads(body)
    content = obj["choices"][0]["message"]["content"]
    return json.loads(content)


def ai_resolve_fuzzy_match(ai: AIConfig, ingredient_name: str, ingredient_cas: str, candidates: List[dict]) -> dict:
    system_prompt = (
        "You are a conservative cosmetic ingredient screening assistant. "
        "Your job is NOT to hallucinate. Only select a candidate if the evidence is strong. "
        "Return strict JSON."
    )
    user_prompt = json.dumps(
        {
            "task": "Decide whether the raw ingredient likely matches one of the database candidates.",
            "rules": [
                "Prefer exact or near-exact INCI/chemical name agreement.",
                "If CAS conflicts or is absent, be conservative.",
                "If not confident, return decision = MANUAL_REVIEW.",
                "Do not invent synonyms not present in candidates.",
            ],
            "raw_ingredient_name": ingredient_name,
            "raw_cas": ingredient_cas,
            "candidates": candidates,
            "output_schema": {
                "decision": "SELECT | MANUAL_REVIEW | NO_MATCH",
                "selected_rank": 1,
                "normalized_name": "string",
                "confidence": "High | Medium | Low",
                "reason": "string"
            },
        },
        ensure_ascii=False,
    )
    return call_openai_compatible_json(ai, system_prompt, user_prompt, temperature=0.0)


def ai_generate_restriction_note(ai: AIConfig, product_family: str, raw_name: str, raw_cas: str, restricted_row: dict, input_pct: Optional[float]) -> str:
    system_prompt = "You draft concise compliance notes for cosmetic ingredient screening. Be conservative and factual. Return JSON only."
    user_prompt = json.dumps(
        {
            "task": "Draft one short compliance note.",
            "product_family": product_family,
            "raw_name": raw_name,
            "raw_cas": raw_cas,
            "input_w_w_pct": input_pct,
            "restricted_entry": restricted_row,
            "requirements": [
                "Mention body site / condition / max concentration / warning only if present.",
                "If product applicability is uncertain, explicitly say review needed.",
                "No legal overstatement.",
            ],
            "output_schema": {"note": "string"},
        },
        ensure_ascii=False,
    )
    data = call_openai_compatible_json(ai, system_prompt, user_prompt, temperature=0.1)
    return str(data.get("note", "")).strip()


# ============================================================
# 文件检查
# ============================================================
def validate_required_files(paths: Dict[str, Path]) -> Dict[str, Path]:
    print_header("步骤 1：检查项目文件夹结构")
    for k in ["root", "input", "database", "template", "output", "logs"]:
        print(f"{k:10s}: {paths[k]}")

    db_found = find_file_by_keywords(paths["database"], DB_FILE_HINTS)
    template_found = find_template_file(paths["template"])
    input_files = list_excel_files(paths["input"])

    print_header("步骤 2：检查 database 文件")
    for k, v in db_found.items():
        print(f"{k:16s}: {v.name if v else '[未找到]'}")

    print_header("步骤 3：检查 template / input")
    print(f"template: {template_found.name if template_found else '[未找到]'}")
    if input_files:
        print("input:")
        for f in input_files:
            print(f"- {f.name}")
    else:
        print("input: [空]")

    missing = [k for k, v in db_found.items() if v is None]
    if missing:
        raise FileNotFoundError("缺少数据库文件: " + ", ".join(missing))
    if template_found is None:
        raise FileNotFoundError("template 文件夹未找到输出模板")
    if not input_files:
        raise FileNotFoundError("input 文件夹未找到客户配方 Excel")

    return {
        "ca_prohibited": db_found["ca_prohibited"],
        "ca_restricted": db_found["ca_restricted"],
        "cir": db_found["cir"],
        "cfr_banned": db_found["cfr_banned"],
        "template": template_found,
    }


# ============================================================
# 数据库标准化
# ============================================================
def load_ca_prohibited(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path)
    df = df.rename(columns=lambda x: str(x).strip())
    first_col = [c for c in df.columns if "Ingredient" in str(c)][0]
    cas_col = [c for c in df.columns if "CAS" in str(c)][0]
    syn_col = [c for c in df.columns if "Synonyms" in str(c)][0]
    df = df[[first_col, cas_col, syn_col]].copy()
    df.columns = ["ingredient_name", "cas_raw", "synonym_raw"]
    df = df[df["ingredient_name"].notna()].copy()
    df = df[df["ingredient_name"].astype(str).str.strip() != ""].copy()
    df["db_source"] = "CA_PROHIBITED"
    df["cas_list"] = df["cas_raw"].apply(lambda x: [normalize_cas(v) for v in split_multi_cas_value(x) if normalize_cas(v)])
    df["synonym_list"] = df["synonym_raw"].apply(lambda x: [normalize_name_key(v) for v in split_multi_text_value(x) if normalize_name_key(v)])
    df["name_key"] = df["ingredient_name"].apply(normalize_name_key)
    df["evidence_text"] = df["ingredient_name"].fillna("").astype(str)
    return df[["db_source", "ingredient_name", "name_key", "cas_list", "synonym_list", "evidence_text"]].reset_index(drop=True)




def _expanded_sheet_values(ws) -> List[List[object]]:
    values = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(1, ws.max_row + 1)]
    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        top_val = ws.cell(min_row, min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                values[r - 1][c - 1] = top_val
    return values


def _join_unique_text(parts: List[object]) -> str:
    out = []
    seen = set()
    for part in parts:
        if part is None or (isinstance(part, float) and pd.isna(part)):
            continue
        s = str(part).replace("\xa0", " ").strip()
        if not s:
            continue
        if s not in seen:
            seen.add(s)
            out.append(s)
    return "\n".join(out)

def load_ca_restricted(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    values = _expanded_sheet_values(ws)

    rows = []
    current = None

    for raw in values:
        chem = raw[0] if len(raw) > 0 else None
        cas = raw[1] if len(raw) > 1 else None
        syn = raw[2] if len(raw) > 2 else None
        cond = raw[3] if len(raw) > 3 else None
        maxc = raw[4] if len(raw) > 4 else None
        warn = raw[5] if len(raw) > 5 else None

        chem_s = "" if chem is None else str(chem).replace("\xa0", " ").strip()
        cas_s = "" if cas is None else str(cas).replace("\xa0", " ").strip()
        syn_s = "" if syn is None else str(syn).replace("\xa0", " ").strip()

        if not chem_s or chem_s.lower() == "chemical":
            continue

        # 对有合并单元格的 restricted 表，最稳妥的分组方式是：
        # 先把 merged cells 展开，再按“同一 ingredient 的连续行”聚合。
        if current is None or current["ingredient_name"] != chem_s:
            if current is not None:
                rows.append(current)
            current = {
                "db_source": "CA_RESTRICTED",
                "ingredient_name": chem_s,
                "name_key": normalize_name_key(chem_s),
                "cas_list": [],
                "synonym_list": [],
                "_conditions_parts": [],
                "_maxc_parts": [],
                "_warn_parts": [],
            }

        for v in split_multi_cas_value(cas_s):
            nv = normalize_cas(v)
            if nv and nv not in current["cas_list"]:
                current["cas_list"].append(nv)
        for v in split_multi_text_value(syn_s):
            nk = normalize_name_key(v)
            if nk and nk not in current["synonym_list"]:
                current["synonym_list"].append(nk)

        current["_conditions_parts"].append(cond)
        current["_maxc_parts"].append(maxc)
        current["_warn_parts"].append(warn)

    if current is not None:
        rows.append(current)

    final_rows = []
    for row in rows:
        final_rows.append({
            "db_source": row["db_source"],
            "ingredient_name": row["ingredient_name"],
            "name_key": row["name_key"],
            "cas_list": row["cas_list"],
            "synonym_list": row["synonym_list"],
            "conditions_of_use": _join_unique_text(row["_conditions_parts"]),
            "max_concentration": _join_unique_text(row["_maxc_parts"]),
            "warning_statement": _join_unique_text(row["_warn_parts"]),
        })
    return pd.DataFrame(final_rows)


def load_cir(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path)
    cols = list(df.columns)
    df = df[[cols[0], cols[1], cols[2]]].copy()
    df.columns = ["ingredient_name", "finding", "conclusion"]
    df = df[df["ingredient_name"].notna()].copy()
    df["ingredient_name"] = df["ingredient_name"].astype(str).str.strip()
    df = df[df["ingredient_name"] != ""].copy()
    df["db_source"] = "CIR"
    df["name_key"] = df["ingredient_name"].apply(normalize_name_key)
    return df[["db_source", "ingredient_name", "name_key", "finding", "conclusion"]].reset_index(drop=True)


def load_cfr_banned(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path)
    first_col = df.columns[0]
    ser = df[first_col].dropna().astype(str).str.strip()
    ser = ser[ser != ""]
    out = pd.DataFrame({"ingredient_name": ser})
    out["db_source"] = "CFR_BANNED"
    out["name_key"] = out["ingredient_name"].apply(normalize_name_key)
    return out[["db_source", "ingredient_name", "name_key"]].reset_index(drop=True)


def build_db_bundle(file_map: Dict[str, Path]) -> Dict[str, pd.DataFrame]:
    return {
        "ca_prohibited": load_ca_prohibited(file_map["ca_prohibited"]),
        "ca_restricted": load_ca_restricted(file_map["ca_restricted"]),
        "cir": load_cir(file_map["cir"]),
        "cfr_banned": load_cfr_banned(file_map["cfr_banned"]),
    }


# ============================================================
# 客户配方解析
# ============================================================
def find_header_row(df_preview: pd.DataFrame) -> int:
    for idx in range(min(len(df_preview), 10)):
        row = [normalize_text(x) for x in df_preview.iloc[idx].tolist()]
        if "no." in row and any("inci" in x for x in row) and any("cas" in x for x in row):
            return idx
    return 0


def parse_client_file(file_path: Path) -> List[dict]:
    records = []
    xls = pd.ExcelFile(file_path)
    for sheet_name in xls.sheet_names:
        raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        if raw.empty:
            continue
        header_row = find_header_row(raw)
        header_values = raw.iloc[header_row].tolist()
        data = raw.iloc[header_row + 1 :].copy()
        data.columns = header_values
        data = data.reset_index(drop=True)

        cols = list(data.columns)
        no_col = next((c for c in cols if normalize_text(c) == "no."), cols[0])
        inci_col = next((c for c in cols if "inci" in normalize_text(c)), cols[1] if len(cols) > 1 else cols[0])
        cas_col = next((c for c in cols if "cas" in normalize_text(c)), cols[2] if len(cols) > 2 else None)
        formula_cols = [c for c in cols if c not in {no_col, inci_col, cas_col}]
        formula_cols = [c for c in formula_cols if str(c).strip() != ""]

        if not formula_cols:
            continue

        base_title = str(raw.iloc[0, 0]).strip() if pd.notna(raw.iloc[0, 0]) else sheet_name
        product_family = sheet_name.strip()

        for fcol in formula_cols:
            fcol_text = str(fcol).replace("\n", " ").strip()
            formula_code = fcol_text.split()[0].strip() if re.search(r"[A-Za-z]{2,}\d+", fcol_text) else re.sub(r"\s+", " ", fcol_text)[:40]
            formula_name = fcol_text
            if " - " not in formula_name and product_family.upper() not in formula_name.upper():
                formula_name = f"{formula_code} - {formula_name}" if formula_code != formula_name else formula_name

            for _, row in data.iterrows():
                raw_inci = row.get(inci_col)
                raw_cas = row.get(cas_col) if cas_col is not None else None
                pct = row.get(fcol)
                ing_no = row.get(no_col)
                if pd.isna(raw_inci) or str(raw_inci).strip() == "":
                    continue
                pct_float = safe_float(pct)
                if pct_float is None:
                    continue
                records.append({
                    "Client File": file_path.name,
                    "Source Sheet": sheet_name,
                    "Formula Code": formula_code,
                    "Formula Name": formula_name,
                    "Product Family": product_family,
                    "Ingredient No": ing_no,
                    "Raw INCI Name": str(raw_inci).strip(),
                    "Raw CAS No": "" if pd.isna(raw_cas) else str(raw_cas).strip(),
                    "Input W/W %": pct_float,
                    "Variant Group": formula_code,
                })
    return records


def load_client_records(input_files: List[Path]) -> pd.DataFrame:
    rows = []
    for f in input_files:
        rows.extend(parse_client_file(f))
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    df.insert(0, "Record ID", range(1, len(df) + 1))
    return df


# ============================================================
# 索引与匹配
# ============================================================
def build_name_index(df: pd.DataFrame, include_synonyms: bool = False) -> Dict[str, List[int]]:
    index: Dict[str, List[int]] = {}
    for i, row in df.iterrows():
        keys = []
        keys.extend(row.get("name_variants", []) or [])
        keys.append(row.get("name_key", ""))
        if include_synonyms and "synonym_list" in df.columns:
            keys.extend(row.get("synonym_list", []) or [])
        for key in keys:
            if not key:
                continue
            index.setdefault(key, []).append(i)
    return index


def build_cas_index(df: pd.DataFrame) -> Dict[str, List[int]]:
    index: Dict[str, List[int]] = {}
    if "cas_list" not in df.columns:
        return index
    for i, row in df.iterrows():
        for cas in row.get("cas_list", []) or []:
            if cas:
                index.setdefault(cas, []).append(i)
    return index


def make_candidate_pool(db_bundle: Dict[str, pd.DataFrame]) -> List[dict]:
    pool = []
    for db_name, df in db_bundle.items():
        for _, row in df.iterrows():
            pool.append({
                "source_db": db_name.upper(),
                "ingredient_name": row.get("ingredient_name", ""),
                "name_key": row.get("name_key", ""),
                "cas": "; ".join(row.get("cas_list", []) or []) if "cas_list" in row else "",
            })
    return pool


def fuzzy_candidates(raw_name: str, candidate_pool: List[dict], topn: int = MAX_FUZZY_CANDIDATES) -> List[dict]:
    key = normalize_name_key(raw_name)
    scored = []
    seen = set()
    for c in candidate_pool:
        reason = low_confidence_name_reason(raw_name, c["ingredient_name"])
        if not reason:
            continue
        uniq = (c["source_db"], c["ingredient_name"])
        if uniq in seen:
            continue
        seen.add(uniq)
        score = fuzz.ratio(key, c["name_key"])
        scored.append({**c, "score": score, "reason": reason})
    scored.sort(key=lambda x: (-x["score"], x["ingredient_name"]))
    return scored[:topn]



def lookup_hits(db_name: str, name_keys: List[str], cas_keys: List[str], db_bundle: Dict[str, pd.DataFrame], indexes: Dict[str, dict], include_synonyms: bool = True) -> List[dict]:
    hits: List[dict] = []
    seen = set()
    for cas in cas_keys:
        for i in indexes[db_name]["cas"].get(cas, []):
            if i not in seen:
                hits.append(db_bundle[db_name].iloc[i].to_dict())
                seen.add(i)
    for key in name_keys:
        for i in indexes[db_name]["name"].get(key, []):
            if i not in seen:
                hits.append(db_bundle[db_name].iloc[i].to_dict())
                seen.add(i)
        if include_synonyms:
            for i in indexes[db_name].get("syn", {}).get(key, []):
                if i not in seen:
                    hits.append(db_bundle[db_name].iloc[i].to_dict())
                    seen.add(i)
    return hits


def first_nonempty(*vals):
    for v in vals:
        if v is None:
            continue
        if isinstance(v, float) and pd.isna(v):
            continue
        if str(v).strip() != "":
            return v
    return ""


def join_hit_ingredients(rows: List[dict]) -> str:
    vals = []
    for r in rows:
        v = str(r.get("ingredient_name", "")).strip()
        if v and v not in vals:
            vals.append(v)
    return " | ".join(vals)


def join_hit_cas(rows: List[dict]) -> str:
    vals = []
    for r in rows:
        for c in r.get("cas_list", []) or []:
            if c and c not in vals:
                vals.append(c)
    return "; ".join(vals)


def dedupe_cir_hit_rows(rows: List[dict]) -> List[dict]:
    out = []
    seen = set()
    for row in rows:
        finding = row.get("finding", "")
        conclusion = row.get("conclusion", "")
        if isinstance(finding, float) and pd.isna(finding):
            finding = ""
        if isinstance(conclusion, float) and pd.isna(conclusion):
            conclusion = ""
        key = (
            str(row.get("ingredient_name", "")).strip(),
            str(finding).strip(),
            str(conclusion).strip(),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out


def sort_cir_hit_rows(rows: List[dict], raw_name_keys: List[str]) -> List[dict]:
    return sorted(
        rows,
        key=lambda r: (
            0 if normalize_name_key(r.get("ingredient_name", "")) in raw_name_keys else 1,
            len(str(r.get("ingredient_name", "")).strip()),
            str(r.get("ingredient_name", "")).strip().lower(),
        ),
    )


def format_cir_field(rows: List[dict], field_name: str) -> str:
    if not rows:
        return ""
    if len(rows) == 1:
        value = rows[0].get(field_name, "")
        if isinstance(value, float) and pd.isna(value):
            return ""
        return str(value or "").strip()

    parts = []
    for row in rows:
        ingredient = str(row.get("ingredient_name", "")).strip()
        value = row.get(field_name, "")
        if isinstance(value, float) and pd.isna(value):
            value = ""
        value = str(value or "").strip() or "(blank in CIR source)"
        parts.append(f"{ingredient}: {value}")
    return "\n---\n".join(parts)


def join_cir_hit_ingredients(rows: List[dict]) -> str:
    vals = []
    for row in rows:
        ingredient = str(row.get("ingredient_name", "")).strip()
        if ingredient:
            vals.append(ingredient)
    return " | ".join(vals)


def build_cir_interpretation(rows: List[dict], raw_name_keys: List[str]) -> str:
    if not rows:
        return ""
    exact_rows = [r for r in rows if normalize_name_key(r.get("ingredient_name", "")) in raw_name_keys]
    if len(rows) == 1:
        return "Exact CIR row matched." if exact_rows else "Related CIR row matched by component/variant name."
    if exact_rows and len(exact_rows) == len(rows):
        return f"Multiple exact CIR rows matched for the same ingredient. All {len(rows)} rows are retained below in order."
    if exact_rows:
        return f"Multiple CIR rows matched. Exact ingredient row plus {len(rows) - len(exact_rows)} related CIR rows are retained below in order."
    return f"Multiple related CIR rows matched by component/variant name. All {len(rows)} CIR rows are retained below in order."


def match_one_record(rec: dict, db_bundle: Dict[str, pd.DataFrame], indexes: Dict[str, dict], candidate_pool: List[dict], ai: AIConfig, fuzzy_log_rows: List[dict]) -> dict:
    raw_name = rec["Raw INCI Name"]
    raw_cas = rec["Raw CAS No"]
    raw_name_keys = name_key_variants(raw_name)
    raw_cas_keys = cas_variants(raw_cas)

    normalized_name = raw_name
    normalized_cas = raw_cas
    match_method = "NO_MATCH"
    match_confidence = "Low"
    manual_review = "N"
    reviewer_notes = []

    out = {
        "Record ID": rec["Record ID"],
        "Formula Code": rec["Formula Code"],
        "Formula Name": rec["Formula Name"],
        "Ingredient No": rec["Ingredient No"],
        "Raw INCI Name": raw_name,
        "Raw CAS No": raw_cas,
        "Input W/W %": rec["Input W/W %"],
        "Product Family": rec["Product Family"],
        "Normalized / Screening Key Name": normalized_name,
        "Normalized / Screening Key CAS": normalized_cas,
        "Match Method": match_method,
        "Match Confidence": match_confidence,
        "Manual Review?": manual_review,
        "CA Prohibited Hit": "N",
        "CA Prohibited Matched Ingredient": "",
        "CA Prohibited CAS": "",
        "CA Prohibited Synonyms / Related": "",
        "CA Prohibited Evidence": "",
        "CA Restricted Hit": "N",
        "CA Restricted Matched Ingredient": "",
        "CA Restricted CAS": "",
        "Conditions of Use / Allowed Body Site": "",
        "Max Conc. Permitted": "",
        "Warning Statement": "",
        "Restriction Assessment Note": "",
        "CIR Hit": "N",
        "CIR Matched Ingredient": "",
        "Finding": "",
        "Conclusion": "",
        "CIR Interpretation": "",
        "CFR Hit": "N",
        "CFR Matched Substance": "",
        "CFR Evidence": "",
        "Hard Stop?": "N",
        "Regulatory Summary": "No flagged hit yet.",
        "Recommended Action": "Proceed with routine review.",
        "Final Disposition": "",
        "Reviewer / AI Notes": "",
        "Checked By": "",
        "Checked Date": "",
    }

    lookup_name_keys = list(raw_name_keys)
    lookup_cas_keys = list(raw_cas_keys)

    # CA prohibited
    ca_pro_hit_rows = lookup_hits("ca_prohibited", lookup_name_keys, lookup_cas_keys, db_bundle, indexes, include_synonyms=True)
    if ca_pro_hit_rows:
        row = ca_pro_hit_rows[0]
        out["CA Prohibited Hit"] = "Y"
        out["CA Prohibited Matched Ingredient"] = join_hit_ingredients(ca_pro_hit_rows)
        out["CA Prohibited CAS"] = join_hit_cas(ca_pro_hit_rows)
        out["CA Prohibited Synonyms / Related"] = "; ".join(row.get("synonym_list", []) or [])
        out["CA Prohibited Evidence"] = row.get("evidence_text", "")

    # CA restricted
    ca_res_hit_rows = lookup_hits("ca_restricted", lookup_name_keys, lookup_cas_keys, db_bundle, indexes, include_synonyms=True)
    if ca_res_hit_rows:
        row = ca_res_hit_rows[0]
        out["CA Restricted Hit"] = "Y"
        out["CA Restricted Matched Ingredient"] = join_hit_ingredients(ca_res_hit_rows)
        out["CA Restricted CAS"] = join_hit_cas(ca_res_hit_rows)
        out["Conditions of Use / Allowed Body Site"] = "\n---\n".join([r.get("conditions_of_use", "") for r in ca_res_hit_rows if str(r.get("conditions_of_use", "")).strip()])
        out["Max Conc. Permitted"] = "\n---\n".join([r.get("max_concentration", "") for r in ca_res_hit_rows if str(r.get("max_concentration", "")).strip()])
        out["Warning Statement"] = "\n---\n".join([r.get("warning_statement", "") for r in ca_res_hit_rows if str(r.get("warning_statement", "")).strip()])
        if ai.enabled:
            try:
                out["Restriction Assessment Note"] = ai_generate_restriction_note(
                    ai, rec["Product Family"], raw_name, raw_cas, row, rec["Input W/W %"],
                )
            except Exception:
                pass
        if not out["Restriction Assessment Note"]:
            bits = []
            if out["Conditions of Use / Allowed Body Site"]:
                bits.append("Check product type/body site applicability")
            if out["Max Conc. Permitted"]:
                bits.append("Compare against input concentration")
            if out["Warning Statement"]:
                bits.append("Verify warning statement on label")
            out["Restriction Assessment Note"] = "; ".join(bits) if bits else "Restricted entry found. Review required."

    # CIR independent lookup
    cir_lookup_name_keys = exact_name_keys(raw_name)
    cir_hit_rows = lookup_hits("cir", cir_lookup_name_keys, [], db_bundle, indexes, include_synonyms=False)
    cir_hit_rows = sort_cir_hit_rows(dedupe_cir_hit_rows(cir_hit_rows), cir_lookup_name_keys)
    if cir_hit_rows:
        out["CIR Hit"] = "Y"
        out["CIR Matched Ingredient"] = join_cir_hit_ingredients(cir_hit_rows)
        out["Finding"] = format_cir_field(cir_hit_rows, "finding")
        out["Conclusion"] = format_cir_field(cir_hit_rows, "conclusion")
        out["CIR Interpretation"] = build_cir_interpretation(cir_hit_rows, cir_lookup_name_keys)
        if len(cir_hit_rows) == 1 and not out["Conclusion"]:
            reviewer_notes.append("CIR matched, but conclusion cell is blank in source table.")
        if len(cir_hit_rows) > 1:
            reviewer_notes.append(f"Multiple CIR rows matched; Finding/Conclusion are written row-by-row for {len(cir_hit_rows)} CIR entries.")

    # CFR
    cfr_hit_rows = lookup_hits("cfr_banned", lookup_name_keys, [], db_bundle, indexes, include_synonyms=False)
    if cfr_hit_rows:
        row = cfr_hit_rows[0]
        out["CFR Hit"] = "Y"
        out["CFR Matched Substance"] = join_hit_ingredients(cfr_hit_rows)
        out["CFR Evidence"] = row.get("ingredient_name", "")

    best_hits = [
        best_row_match_evidence(ca_pro_hit_rows, raw_name_keys, raw_cas_keys),
        best_row_match_evidence(ca_res_hit_rows, raw_name_keys, raw_cas_keys),
        best_row_match_evidence(cir_hit_rows, raw_name_keys, raw_cas_keys),
        best_row_match_evidence(cfr_hit_rows, raw_name_keys, raw_cas_keys),
    ]
    best_hits = [x for x in best_hits if x is not None]
    best_match = best_hits[0] if best_hits else None
    if best_hits:
        best_match = sorted(
            best_hits,
            key=lambda x: (
                {"NAME_AND_CAS_MATCH": 0, "CAS_MATCH_ONLY": 1, "NAME_MATCH_CAS_MISMATCH": 2, "NAME_MATCH_ONLY": 3}.get(x["category"], 99),
                {"CA_PROHIBITED": 0, "CA_RESTRICTED": 1, "CFR_BANNED": 2, "CIR": 3}.get(x["source_db"], 99),
                0 if x["name_via_primary"] else 1,
            ),
        )[0]

    if best_match:
        matched_row = best_match["row"]
        if best_match["name_via_primary"]:
            normalized_name = matched_row.get("ingredient_name", normalized_name) or normalized_name
        if best_match["cas_match"] and raw_cas_keys:
            normalized_cas = raw_cas_keys[0]

        if best_match["category"] == "NAME_AND_CAS_MATCH":
            match_method = "NAME_AND_CAS_MATCH"
            match_confidence = "High"
        elif best_match["category"] == "CAS_MATCH_ONLY":
            match_method = "CAS_MATCH_NAME_DIFFER"
            match_confidence = "High"
            manual_review = "Y"
        elif best_match["category"] == "NAME_MATCH_CAS_MISMATCH":
            match_method = "NAME_MATCH_CAS_MISMATCH"
            match_confidence = "Medium"
            manual_review = "Y"
        else:
            match_method = "NAME_MATCH_ONLY"
            match_confidence = "High" if best_match["name_via_primary"] else "Medium"

        if best_match["name_via_synonym"] and not best_match["name_via_primary"]:
            reviewer_notes.append(f"{best_match['source_db']} matched via database synonym.")
    else:
        candidates = fuzzy_candidates(raw_name, candidate_pool, MAX_FUZZY_CANDIDATES)
        for rank, c in enumerate(candidates, start=1):
            fuzzy_log_rows.append({
                "Record ID": rec["Record ID"],
                "Formula Code": rec["Formula Code"],
                "Raw INCI Name": raw_name,
                "Raw CAS No": raw_cas,
                "Candidate Rank": rank,
                "Candidate Ingredient": c["ingredient_name"],
                "Candidate Source DB": c["source_db"],
                "Candidate CAS": c["cas"],
                "Similarity Score": c["score"],
                "Why Suggested": c.get("reason", "Low-confidence near match"),
                "Final Selected?": "",
                "Reviewer Note": "",
            })
        if candidates:
            top = candidates[0]
            match_method = "LOW_CONFIDENCE_NAME_MATCH"
            match_confidence = "Low"
            manual_review = "Y"
            reviewer_notes.append(
                f"Possible near-name match in {top['source_db']}: {top['ingredient_name']} ({top.get('reason', 'low-confidence similarity')})."
            )

    ca_pro_best = best_row_match_evidence(ca_pro_hit_rows, raw_name_keys, raw_cas_keys)
    if ca_pro_best and ca_pro_best["category"] == "NAME_MATCH_CAS_MISMATCH":
        manual_review = "Y"
        reviewer_notes.append("CA Prohibited hit needs manual review because name matched but CAS did not.")
    if ca_pro_best and ca_pro_best["category"] == "CAS_MATCH_ONLY":
        manual_review = "Y"
        reviewer_notes.append("CA Prohibited hit needs manual review because CAS matched but name did not.")

    ca_res_best = best_row_match_evidence(ca_res_hit_rows, raw_name_keys, raw_cas_keys)
    if ca_res_best and ca_res_best["category"] == "NAME_MATCH_CAS_MISMATCH":
        manual_review = "Y"
        reviewer_notes.append("CA Restricted hit needs manual review because name matched but CAS did not.")
    if ca_res_best and ca_res_best["category"] == "CAS_MATCH_ONLY":
        manual_review = "Y"
        reviewer_notes.append("CA Restricted hit needs manual review because CAS matched but name did not.")
    if len(ca_res_hit_rows) > 1:
        manual_review = "Y"
        reviewer_notes.append("Multiple CA Restricted entries matched; verify which restriction line applies.")

    out["Manual Review?"] = manual_review
    out["Normalized / Screening Key Name"] = normalized_name
    out["Normalized / Screening Key CAS"] = normalized_cas
    out["Match Method"] = match_method
    out["Match Confidence"] = match_confidence

    if out["CA Prohibited Hit"] == "Y" or out["CFR Hit"] == "Y":
        out["Hard Stop?"] = "Y"
        out["Regulatory Summary"] = "Prohibited hit identified in CA Hotlist and/or CFR."
        out["Recommended Action"] = "Escalate immediately before commercialization."
    elif out["CA Restricted Hit"] == "Y":
        out["Hard Stop?"] = "REVIEW"
        out["Regulatory Summary"] = "Restricted substance identified. Product type / concentration / warning review needed."
        out["Recommended Action"] = "Check body site, max concentration, and warning statement applicability."
    elif out["Manual Review?"] in {"Y", "MAYBE"}:
        out["Hard Stop?"] = "REVIEW"
        out["Regulatory Summary"] = "No confirmed prohibited/restricted hit, but identity / normalization still needs review."
        out["Recommended Action"] = "Resolve fuzzy match / identity issue before final conclusion."
    elif out["CIR Hit"] == "Y":
        out["Hard Stop?"] = "N"
        out["Regulatory Summary"] = "CIR reference found only."
        out["Recommended Action"] = "Proceed with routine review and note CIR finding/conclusion."
    else:
        out["Hard Stop?"] = "N"
        out["Regulatory Summary"] = "No CA/CFR prohibition or restriction hit; no CIR reference found."
        out["Recommended Action"] = "No manual review required unless there is external context or labeling concern."

    out["Reviewer / AI Notes"] = " | ".join([x for x in reviewer_notes if x])
    return out

def run_screening(client_df: pd.DataFrame, db_bundle: Dict[str, pd.DataFrame], ai: AIConfig) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    indexes = {}
    for dbk, df in db_bundle.items():
        indexes[dbk] = {
            "name": build_name_index(df, include_synonyms=False),
            "syn": build_name_index(df, include_synonyms=True) if "synonym_list" in df.columns else {},
            "cas": build_cas_index(df),
        }
    candidate_pool = make_candidate_pool(db_bundle)
    fuzzy_log_rows = []
    output_rows = []
    for _, rec in client_df.iterrows():
        output_rows.append(match_one_record(rec.to_dict(), db_bundle, indexes, candidate_pool, ai, fuzzy_log_rows))
    output_df = pd.DataFrame(output_rows)
    fuzzy_df = pd.DataFrame(fuzzy_log_rows)

    summary_rows = []
    if not output_df.empty:
        for formula_code, grp in output_df.groupby("Formula Code", dropna=False):
            summary_rows.append({
                "Formula Code": formula_code,
                "Formula Name": grp["Formula Name"].iloc[0],
                "Product Family": grp["Product Family"].iloc[0],
                "Ingredient Count": len(grp),
                "CA Prohibited Hits": int((grp["CA Prohibited Hit"] == "Y").sum()),
                "CA Restricted Hits": int((grp["CA Restricted Hit"] == "Y").sum()),
                "CIR References": int((grp["CIR Hit"] == "Y").sum()),
                "CFR Hits": int((grp["CFR Hit"] == "Y").sum()),
                "Manual Review Count": int(grp["Manual Review?"].isin(["Y", "MAYBE"]).sum()),
                "Overall Formula Status": (
                    "Prohibited hit present" if ((grp["CA Prohibited Hit"] == "Y").any() or (grp["CFR Hit"] == "Y").any())
                    else "Restricted review needed" if (grp["CA Restricted Hit"] == "Y").any()
                    else "Manual review needed" if grp["Manual Review?"].isin(["Y", "MAYBE"]).any()
                    else "No flagged hits yet"
                ),
                "Follow-up": (
                    "Escalate immediately" if ((grp["CA Prohibited Hit"] == "Y").any() or (grp["CFR Hit"] == "Y").any())
                    else "Check body site / max conc. / warnings" if (grp["CA Restricted Hit"] == "Y").any()
                    else "Resolve fuzzy match / naming" if grp["Manual Review?"].isin(["Y", "MAYBE"]).any()
                    else "Proceed with routine review"
                ),
            })
    summary_df = pd.DataFrame(summary_rows)
    return output_df, summary_df, fuzzy_df


# ============================================================
# 写回模板
# ============================================================
def clear_sheet_body(ws, start_row: int) -> None:
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(start_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).value = None


def write_df_to_sheet(ws, df: pd.DataFrame, start_row: int, headers: List[str]) -> None:
    if df.empty:
        return
    for row_idx, (_, rec) in enumerate(df.iterrows(), start=start_row):
        for col_idx, h in enumerate(headers, start=1):
            val = rec.get(h, "")
            ws.cell(row_idx, col_idx).value = val


def make_unique_sheet_title(base_title: str, existing_titles: set[str]) -> str:
    invalid_chars = set(r'[]:*?/\\')
    clean = "".join("_" if ch in invalid_chars else ch for ch in str(base_title).strip())
    clean = clean.strip("'") or "Sheet"
    max_len = 31
    title = clean[:max_len]
    if title not in existing_titles:
        return title

    counter = 2
    while True:
        suffix = f"_{counter}"
        candidate = clean[: max_len - len(suffix)] + suffix
        if candidate not in existing_titles:
            return candidate
        counter += 1


def style_output_sheet(ws, row_count: int) -> None:
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    col_map = {str(ws.cell(2, c).value).strip(): c for c in range(1, ws.max_column + 1)}

    hidden_headers = [
        "CA Prohibited Matched Ingredient",
        "CA Prohibited CAS",
        "CA Prohibited Synonyms / Related",
        "CA Prohibited Evidence",
        "CFR Matched Substance",
        "CFR Evidence",
    ]
    for header in hidden_headers:
        col_idx = col_map.get(header)
        if col_idx:
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    red_fill = PatternFill(fill_type="solid", fgColor="FDE9E7")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    orange_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    green_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")

    prohibited_col = col_map.get("CA Prohibited Hit")
    restricted_col = col_map.get("CA Restricted Hit")
    cir_col = col_map.get("CIR Hit")
    cfr_col = col_map.get("CFR Hit")
    manual_col = col_map.get("Manual Review?")

    for r in range(3, 3 + row_count):
        prohibited_val = str(ws.cell(r, prohibited_col).value).strip() if prohibited_col else ""
        restricted_val = str(ws.cell(r, restricted_col).value).strip() if restricted_col else ""
        cir_val = str(ws.cell(r, cir_col).value).strip() if cir_col else ""
        cfr_val = str(ws.cell(r, cfr_col).value).strip() if cfr_col else ""
        manual_val = str(ws.cell(r, manual_col).value).strip() if manual_col else ""

        row_fill = white_fill
        if prohibited_val == "Y" or cfr_val == "Y":
            row_fill = red_fill
        elif restricted_val == "Y":
            row_fill = yellow_fill
        elif manual_val in {"Y", "MAYBE"}:
            row_fill = orange_fill

        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).fill = row_fill

        if (
            cir_col
            and cir_val == "Y"
            and prohibited_val != "Y"
            and cfr_val != "Y"
            and restricted_val != "Y"
            and manual_val not in {"Y", "MAYBE"}
        ):
            ws.cell(r, cir_col).fill = green_fill



def build_output_workbook(template_path: Path, output_path: Path, client_df: pd.DataFrame, output_df: pd.DataFrame, summary_df: pd.DataFrame, fuzzy_df: pd.DataFrame) -> None:
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)

    ws02 = wb["02_客户成分汇总"]
    ws03 = wb["03_筛查输出模板"]
    ws04 = wb["04_配方汇总"]
    ws05 = wb["05_模糊匹配日志"]

    clear_sheet_body(ws02, 2)
    clear_sheet_body(ws03, 3)
    clear_sheet_body(ws04, 2)
    clear_sheet_body(ws05, 2)

    headers02 = [ws02.cell(1, c).value for c in range(1, ws02.max_column + 1)]
    headers03 = [ws03.cell(2, c).value for c in range(1, ws03.max_column + 1)]
    headers04 = [ws04.cell(1, c).value for c in range(1, ws04.max_column + 1)]
    headers05 = [ws05.cell(1, c).value for c in range(1, ws05.max_column + 1)]

    write_df_to_sheet(ws02, client_df, 2, headers02)
    write_df_to_sheet(ws04, summary_df, 2, headers04)
    write_df_to_sheet(ws05, fuzzy_df, 2, headers05)
    existing_titles = {ws.title for ws in wb.worksheets}
    formula_groups = output_df.groupby("Formula Code", dropna=False, sort=False) if not output_df.empty else []
    for formula_code, grp in formula_groups:
        sheet_title = make_unique_sheet_title(f"03_{formula_code}", existing_titles)
        existing_titles.add(sheet_title)
        ws_formula = wb.copy_worksheet(ws03)
        ws_formula.title = sheet_title
        clear_sheet_body(ws_formula, 3)
        write_df_to_sheet(ws_formula, grp, 3, headers03)
        style_output_sheet(ws_formula, len(grp))

    ws03.sheet_state = "hidden"

    wb.save(output_path)

def write_run_log(log_path: Path, file_map: Dict[str, Path], ai: AIConfig, counts: Dict[str, int]) -> None:
    lines = [
        f"Run time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"AI enabled: {ai.enabled}",
        f"Model: {ai.model if ai.enabled else '[disabled]'}",
        "Files:",
    ]
    for k, v in file_map.items():
        lines.append(f"- {k}: {v}")
    lines.append("Counts:")
    for k, v in counts.items():
        lines.append(f"- {k}: {v}")
    log_path.write_text("\n".join(lines), encoding="utf-8")


# ============================================================
# 主流程
# ============================================================
def main() -> None:
    root = project_root()
    paths = ensure_project_structure(root)
    ai = load_ai_config(root)

    print_header("CA & US 成分审核工具（Windows + Python + AI）")
    print("项目目录:", root)
    print("AI模式:", "开启" if ai.enabled else "未开启（将仅使用规则匹配）")
    if not ai.enabled and AI_MODE:
        print("提示: 当前未检测到 API Key。请打开本脚本顶部，把 DASHSCOPE_API_KEY 填进去。")

    file_map = validate_required_files(paths)
    input_files = list_excel_files(paths["input"])

    print_header("步骤 4：读取数据库")
    db_bundle = build_db_bundle(file_map)
    for k, df in db_bundle.items():
        print(f"{k:16s}: {len(df)} rows")

    print_header("步骤 5：解析客户输入")
    client_df = load_client_records(input_files)
    print(f"客户成分记录数: {len(client_df)}")
    if client_df.empty:
        raise RuntimeError("未解析出任何客户成分记录，请检查 input 文件格式。")

    print_header("步骤 6：执行筛查")
    output_df, summary_df, fuzzy_df = run_screening(client_df, db_bundle, ai)
    print(f"筛查输出行数: {len(output_df)}")
    print(f"配方汇总行数: {len(summary_df)}")
    print(f"模糊匹配日志行数: {len(fuzzy_df)}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = paths["output"] / f"Cosmetic_Compliance_Output_{ts}.xlsx"
    build_output_workbook(file_map["template"], output_path, client_df, output_df, summary_df, fuzzy_df)

    log_path = paths["logs"] / f"run_log_{ts}.txt"
    write_run_log(log_path, file_map, ai, {
        "client_records": len(client_df),
        "output_rows": len(output_df),
        "summary_rows": len(summary_df),
        "fuzzy_rows": len(fuzzy_df),
    })

    print_header("运行完成")
    print(f"输出文件: {output_path}")
    print(f"日志文件: {log_path}")
    print("\n按回车键退出...")
    input()


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("\n[程序报错]")
        print(str(e))
        print("\n详细报错如下：")
        traceback.print_exc()
        print("\n按回车键退出...")
        try:
            input()
        except Exception:
            pass
        sys.exit(1)
